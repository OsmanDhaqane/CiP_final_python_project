import openpyxl

def load_data(excel_file):
    # Initialize an empty dictionary to store account balances
    accounts = {}
    # Initialize a variable to store the total balance
    total_balance = 0
    try:
        # Load the workbook from the specified Excel file
        wb = openpyxl.load_workbook(excel_file)
        # Get the active sheet of the workbook
        sheet = wb.active
        # Iterate through each row of the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True): # Skip the first row (header row)
            # Extract the name and balance from the current row
            name, balance = row
            # Add the name and balance to the accounts dictionary
            accounts[name] = float(balance)
            # Add the balance to the total balance
            total_balance += float(balance) 
        # Close the workbook
        wb.close()
    except FileNotFoundError:
        # Print an error message if the Excel file is not found
        print("Excel file not found.")
    # Return the accounts dictionary and the total balance
    return accounts, total_balance

def save_data(accounts, excel_file):
    # Create a new Workbook object
    wb = openpyxl.Workbook()
    # Get the active sheet of the workbook
    sheet = wb.active
    # Add headers to the first row
    sheet.append(["Accounts", "Balance"])
    # Iterate through each key-value pair in the accounts dictionary
    for name, balance in accounts.items():
        # Append the name and balance as a row in the sheet
        sheet.append([name, balance])
    # Save the workbook to the specified Excel file
    wb.save(excel_file)

def deposit(accounts, name, amount, excel_file):
    # Check if the account exists in the accounts dictionary
    if name in accounts:
        # If the account exists, add the deposited amount to its current balance
        accounts[name] += amount
    else:
        # If the account doesn't exist, create a new entry with the deposited amount
        accounts[name] = amount
    # After updating the account information, save the data to the Excel file
    save_data(accounts, excel_file)

def withdraw(accounts, name, amount, excel_file):
    # Check if the account exists in the accounts dictionary
    if name in accounts:
        # Check if the account has sufficient funds for withdrawal
        if accounts[name] >= amount:
            # If there are sufficient funds, deduct the withdrawal amount from the account balance
            accounts[name] -= amount
            # Save the updated account data to the Excel file
            save_data(accounts, excel_file)
            # Return True to indicate a successful withdrawal
            return True
        else:
            # If there are insufficient funds, print a message and return False
            print("Insufficient funds!")
            return False
    else:
        # If the account does not exist, print a message and return False
        print("Account not found!")
        return False

def get_balance(accounts, name):
    # Check if the account exists in the accounts dictionary
    if name in accounts:
        # If the account exists, return its balance
        return accounts[name]
    else:
        # If the account does not exist, print a message and return None
        print("Account not found!")
        return None

def get_total_balance(accounts):
    # Calculate the sum of all balances stored in the 'accounts' dictionary
    return sum(accounts.values())


def main():
    # Specify the Excel file containing resident accounts
    excel_file = "resident_accounts.xlsx"
    
    # Load existing account data from the Excel file
    accounts, total_balance = load_data(excel_file)

    # Deposit money into resident accounts
    deposit(accounts, "Ole", 600, excel_file)
    deposit(accounts, "Osman", 500, excel_file)
    deposit(accounts, "Kari", 200, excel_file)
    deposit(accounts, "Ali", 500, excel_file)
    deposit(accounts, "John", 450, excel_file)

    # Withdraw money from resident accounts
    withdraw(accounts, "Ole", 6000, excel_file)
    withdraw(accounts, "Osman", 300, excel_file)

    save_data(accounts, excel_file)

    # Print the total balance across all resident accounts
    print("Total balance:", get_total_balance(accounts))

if __name__ == "__main__":
    main()